"""
report.py — Generate Excel compliance reports from batch match results.
"""

import json
from io import BytesIO
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATUS_COLORS = {
    "match":               "C6EFCE",  # green
    "mismatch":            "FFCCCC",  # red
    "missing_application": "FFF2CC",  # yellow
    "missing_label":       "FFF2CC",  # yellow
    "missing_both":        "F4CCCC",  # darker red
    "error":               "E0E0E0",  # grey
}

OVERALL_COLORS = {
    "match":    "375623",
    "mismatch": "9C0006",
    "error":    "595959",
}

FIELD_COLUMNS = [
    ("brand_name",        "Brand Name"),
    ("class_type",        "Class/Type"),
    ("alcohol_content",   "Alcohol %"),
    ("net_contents",      "Net Contents"),
    ("producer_address",  "Producer/Bottler"),
    ("country_of_origin", "Country of Origin"),
    ("health_warning",    "Health Warning"),
]


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def generate_excel_report(batch_id: str, results: List[Dict], batch_summary: Dict) -> bytes:
    """
    Build a formatted Excel workbook with:
    - Sheet 1: Summary dashboard
    - Sheet 2: Full results per pair
    - Sheet 3: Discrepancies only
    Returns the workbook as bytes.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    sub_fill    = PatternFill("solid", fgColor="2F5496")
    sub_font    = Font(bold=True, color="FFFFFF")

    ws_sum["A1"] = "Alcohol Label Compliance Report"
    ws_sum["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws_sum["A2"] = f"Batch ID: {batch_id}"
    ws_sum["A2"].font = Font(italic=True, color="595959")

    rows = [
        ("Metric", "Value"),
        ("Total Pairs Processed", batch_summary.get("total_pairs", 0)),
        ("Fully Matched",         batch_summary.get("matched", 0)),
        ("Mismatched / Issues",   batch_summary.get("mismatched", 0)),
        ("Processing Errors",     batch_summary.get("errors", 0)),
        ("Overall Pass Rate",     f"{batch_summary.get('pass_rate', 0):.1f}%"),
    ]
    for r_idx, (label, value) in enumerate(rows, start=4):
        cell_a = ws_sum.cell(row=r_idx, column=1, value=label)
        cell_b = ws_sum.cell(row=r_idx, column=2, value=value)
        if r_idx == 4:
            cell_a.fill = header_fill; cell_a.font = header_font
            cell_b.fill = header_fill; cell_b.font = header_font
        else:
            cell_a.font = Font(bold=True)
        cell_a.border = _thin_border()
        cell_b.border = _thin_border()

    # Field-level summary
    ws_sum["A11"] = "Field-Level Match Summary"
    ws_sum["A11"].font = Font(bold=True, size=12, color="1F3864")
    field_headers = ["Field", "Matched", "Mismatched", "Missing (App)", "Missing (Label)", "Missing (Both)"]
    for c, h in enumerate(field_headers, 1):
        cell = ws_sum.cell(row=12, column=c, value=h)
        cell.fill = sub_fill; cell.font = sub_font
        cell.border = _thin_border()
        ws_sum.column_dimensions[get_column_letter(c)].width = 18

    field_names = [f for f, _ in FIELD_COLUMNS]
    for r_idx, (fname, flabel) in enumerate(FIELD_COLUMNS, start=13):
        counts = {"match": 0, "mismatch": 0, "missing_application": 0, "missing_label": 0, "missing_both": 0}
        for res in results:
            status = (res.get("fields") or {}).get(fname, "error")
            if status in counts:
                counts[status] += 1
        ws_sum.cell(row=r_idx, column=1, value=flabel).border = _thin_border()
        for c, key in enumerate(["match", "mismatch", "missing_application", "missing_label", "missing_both"], 2):
            cell = ws_sum.cell(row=r_idx, column=c, value=counts[key])
            cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(key, "FFFFFF"))
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 2: Full Results ─────────────────────────────────────
    ws_res = wb.create_sheet("Full Results")
    headers = ["Pair #", "Overall Status", "Confidence"] + \
              [lbl for _, lbl in FIELD_COLUMNS] + \
              ["Discrepancies", "Notes"]

    for c, h in enumerate(headers, 1):
        cell = ws_res.cell(row=1, column=c, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.border = _thin_border()
        ws_res.column_dimensions[get_column_letter(c)].width = 20

    ws_res.column_dimensions["A"].width = 8
    ws_res.column_dimensions["B"].width = 16
    ws_res.column_dimensions["C"].width = 12

    for r_idx, res in enumerate(results, start=2):
        overall = res.get("overall_status", "error")
        ws_res.cell(row=r_idx, column=1, value=res.get("pair_index", r_idx - 1)).border = _thin_border()

        cell_status = ws_res.cell(row=r_idx, column=2, value=overall.upper())
        cell_status.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(overall, "E0E0E0"))
        cell_status.font = Font(bold=True, color=OVERALL_COLORS.get(overall, "000000"))
        cell_status.border = _thin_border()

        ws_res.cell(row=r_idx, column=3, value=f"{res.get('confidence_score', 0):.0%}").border = _thin_border()

        fields_status = res.get("fields", {})
        for c_offset, (fname, _) in enumerate(FIELD_COLUMNS, start=4):
            status = fields_status.get(fname, "error")
            cell = ws_res.cell(row=r_idx, column=c_offset, value=status)
            cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")

        disc = "; ".join(res.get("discrepancies", []))
        col_disc = len(FIELD_COLUMNS) + 4
        ws_res.cell(row=r_idx, column=col_disc, value=disc).border = _thin_border()
        ws_res.cell(row=r_idx, column=col_disc + 1, value=res.get("notes", "")).border = _thin_border()

    ws_res.column_dimensions[get_column_letter(col_disc)].width = 50
    ws_res.column_dimensions[get_column_letter(col_disc + 1)].width = 30
    ws_res.freeze_panes = "A2"

    # ── Sheet 3: Discrepancies Only ───────────────────────────────
    ws_disc = wb.create_sheet("Discrepancies")
    disc_headers = ["Pair #", "Overall Status", "Field", "Issue", "Discrepancy Detail"]
    for c, h in enumerate(disc_headers, 1):
        cell = ws_disc.cell(row=1, column=c, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.border = _thin_border()
        ws_disc.column_dimensions[get_column_letter(c)].width = [8, 16, 20, 20, 50][c - 1]

    row_num = 2
    for res in results:
        overall = res.get("overall_status", "error")
        if overall == "match":
            continue
        fields_status = res.get("fields", {})
        discrepancies = res.get("discrepancies", [])
        for fname, flabel in FIELD_COLUMNS:
            status = fields_status.get(fname, "match")
            if status != "match":
                ws_disc.cell(row=row_num, column=1, value=res.get("pair_index")).border = _thin_border()
                c_status = ws_disc.cell(row=row_num, column=2, value=overall.upper())
                c_status.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(overall, "FFFFFF"))
                c_status.border = _thin_border()
                ws_disc.cell(row=row_num, column=3, value=flabel).border = _thin_border()
                c_issue = ws_disc.cell(row=row_num, column=4, value=status)
                c_issue.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
                c_issue.border = _thin_border()
                detail = next((d for d in discrepancies if flabel.lower() in d.lower() or fname in d.lower()), "")
                ws_disc.cell(row=row_num, column=5, value=detail).border = _thin_border()
                row_num += 1

    ws_disc.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
